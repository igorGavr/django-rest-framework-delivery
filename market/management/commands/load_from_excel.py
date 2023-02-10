from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from delivery.settings import DATA_DIR
from market.models import Category, Product

class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print(f'Start importing from excel - {DATA_DIR}')
        wb = load_workbook(DATA_DIR+'\price.xlsx')
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        cat = None
        for cnt in range(1, sheet.max_row+1):
            item = sheet.cell(row=cnt, column=2).value
            id = sheet.cell(row=cnt, column=1).value

            if id == None:
                print('Create a new category')
                cat = Category()
                cat.name = item
                cat.save()
            else:
                print('Create a new good')
                if cat:
                    product = Product()
                    product.name = item
                    product.category = cat
                    product.save()